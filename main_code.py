import streamlit as st
import pandas as pd
from collections import Counter
import io
import datetime
import os
import re
from openpyxl import load_workbook

# ==========================================
# 1. PAGE CONFIG & THEME
# ==========================================
st.set_page_config(page_title="Logistics Portal", layout="wide")

st.markdown("""
    <style>
    .main { background-color: #0E1117; }
    div.stButton > button:first-child { 
        background-color: #262730; 
        color: white; 
        border-radius: 5px; 
        width: 100%;
        height: 50px;
        font-size: 20px;
        font-weight: bold;
    }
    .stTable { border: 1px solid #262730; border-radius: 5px; }
    </style>
    """, unsafe_allow_html=True)

# Initialize Session State for Navigation
if 'active_tool' not in st.session_state:
    st.session_state.active_tool = None

# ==========================================
# 2. HELPER FUNCTIONS
# ==========================================
def clean_numeric(value):
    if pd.isna(value) or value == "": return 0.0
    if isinstance(value, (int, float)): return float(value)
    clean_val = str(value).replace(',', '').replace('$', '').strip()
    try: return float(clean_val)
    except ValueError: return 0.0

def clean_sku(val):
    if pd.isna(val): return ""
    s = str(val).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def get_hts_data():
    try:
        current_folder = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_folder, "Cleaned_HTS_Codes.csv")
        df = pd.read_csv(file_path, dtype=str)
        mapping = {}
        for _, row in df.iterrows():
            sku = clean_sku(row.get('SKU', ''))
            if sku:
                mapping[sku] = {
                    "hts": clean_sku(row.get('HTS', '')),
                    "desc": str(row.get('Description', '')).strip()
                }
        return mapping
    except: return {}

def update_detailed_state():
    if "detailed_editor" in st.session_state:
        edits = st.session_state["detailed_editor"]["edited_rows"]
        for row_idx, changes in edits.items():
            for col_name, new_val in changes.items():
                st.session_state.df_detailed.at[row_idx, col_name] = new_val
            q  = st.session_state.df_detailed.at[row_idx, "Quantity"]
            p  = st.session_state.df_detailed.at[row_idx, "Unit Price"]
            uw = st.session_state.df_detailed.at[row_idx, "Unit_Weight_KG"]
            st.session_state.df_detailed.at[row_idx, "Total"] = round(q * p, 2)
            st.session_state.df_detailed.at[row_idx, "Total Weight (KG)"] = round(q * uw, 2)

# ── Destination → template filename mapping ────────────────────────────────
INVOICE_DESTINATIONS = {
    "🇦🇺 Australia / APAC": "Commercial_Invoice_APAC.xlsx",
    "🇪🇺 Europe (EU)":       "Commercial_Invoice_EU.xlsx",
    "🇬🇧 United Kingdom":    "Commercial_Invoice_UK.xlsx",
    "🇨🇦 Canada":            "Commercial_Invoice_CAN.xlsx",
    "🇲🇽 Mexico":            "Commercial_Invoice_MEX.xlsx",
}

# ── Row layout constants (confirmed across all templates) ──────────────────
CI_DATA_START   = 17
CI_DATA_END     = 54
CI_MAX_ROWS     = CI_DATA_END - CI_DATA_START + 1   # 38
CI_SUBTOTAL_ROW = 55
CI_TOTAL_ROW    = 58

def fill_commercial_invoice_template(df_detailed, template_filename):
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), template_filename)
    if not os.path.exists(template_path):
        return None, f"Template file **'{template_filename}'** not found next to main_code.py."
    try:
        wb = load_workbook(template_path)
    except Exception as e:
        return None, f"Could not open template '{template_filename}': {e}"

    ws = wb.active
    po_list = ", ".join(str(p) for p in df_detailed['PO#'].dropna().unique())
    ws['F4'] = po_list
    ws['F5'] = datetime.date.today().strftime("%m/%d/%Y")

    items = df_detailed[['SKU', 'HTS Code', 'Origin', 'Description',
                          'Quantity', 'Unit Price']].values.tolist()
    for i, item in enumerate(items):
        r = CI_DATA_START + i
        ws.cell(row=r, column=1).value = str(item[0])
        ws.cell(row=r, column=2).value = str(item[1])
        ws.cell(row=r, column=3).value = str(item[2])
        ws.cell(row=r, column=4).value = str(item[3])
        ws.cell(row=r, column=5).value = int(item[4])
        ws.cell(row=r, column=6).value = round(float(item[5]), 3)
        ws.cell(row=r, column=7).value = f'=F{r}*E{r}'

    last_data_row = CI_DATA_START + len(items) - 1
    ws[f'E{CI_SUBTOTAL_ROW}'] = f'=SUM(E{CI_DATA_START}:E{last_data_row})'
    ws[f'G{CI_SUBTOTAL_ROW}'] = f'=SUM(G{CI_DATA_START}:G{last_data_row})'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None

def extract_pl_weights_kg(pl_df):
    net_kg = 0.0
    gross_kg = 0.0
    for r in range(len(pl_df) - 1, -1, -1):
        for c in range(len(pl_df.columns)):
            cell_val = str(pl_df.iloc[r, c]).strip().lower()
            if cell_val == "net weight" and net_kg == 0.0:
                try:
                    net_kg = round(clean_numeric(pl_df.iloc[r - 1, c]) * 0.453592, 2)
                except: pass
            if cell_val == "gross weight" and gross_kg == 0.0:
                try:
                    gross_kg = round(clean_numeric(pl_df.iloc[r - 1, c]) * 0.453592, 2)
                except: pass
    return net_kg, gross_kg

def fill_vgw_template(container_num, seal_num, tare_weight, net_kg, gross_kg):
    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "VERIFIED_GROSS_WEIGHT_DECLARATION.xlsx"
    )
    if not os.path.exists(template_path):
        return None, "Template file **'VERIFIED_GROSS_WEIGHT_DECLARATION.xlsx'** not found next to main_code.py."
    try:
        wb = load_workbook(template_path)
    except Exception as e:
        return None, f"Could not open VGW template: {e}"

    ws = wb.active
    ws['B12'] = container_num
    ws['D12'] = seal_num
    ws['B18'] = net_kg
    ws['B20'] = gross_kg
    ws['B22'] = float(tare_weight)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None

def fill_packing_declaration_template(vessel, voyage, consignment, printed_name):
    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "Packing_Declaration_MONAT_AUS.xlsx"
    )
    if not os.path.exists(template_path):
        return None, "Template file **'Packing_Declaration_MONAT_AUS.xlsx'** not found next to main_code.py."
    try:
        wb = load_workbook(template_path)
    except Exception as e:
        return None, f"Could not open Packing Declaration template: {e}"

    ws = wb.active
    ws['D12'] = vessel        # Vessel name
    ws['J12'] = voyage        # Voyage number
    ws['F14'] = consignment   # Consignment identifier (anchor of merged F14:M14)
    ws['G38'] = printed_name  # Printed name
    # Date of issue (A42) is =TODAY() in template — no touch needed

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None

def fill_ausfta_template(bol, name_printed, po_number):
    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "AUSFTA_Declaration_MONAT.xlsx"
    )
    if not os.path.exists(template_path):
        return None, "Template file **'AUSFTA_Declaration_MONAT.xlsx'** not found next to main_code.py."
    try:
        wb = load_workbook(template_path)
    except Exception as e:
        return None, f"Could not open AUSFTA template: {e}"

    ws = wb.active
    ws['D11'] = str(po_number)   # Invoice Number (from PO#)
    ws['D12'] = bol              # Bill of Lading / Airway Bill (manual)
    ws['D13'] = str(po_number)   # Document Reference (same PO#)
    ws['C28'] = name_printed     # Name Printed (manual)
    # C27 (Date) and C29 (Company Name) are already in template — no touch needed

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), None

# ==========================================
# 3. DASHBOARD / TOOL SELECTION (CENTER)
# ==========================================
if st.session_state.active_tool is None:
    st.title("📂 Logistics Operations Portal")
    st.subheader("Select a tool to begin:")
    st.write("---")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("📦 Quote Generator"):
            st.session_state.active_tool = "Quote Pipeline"
            st.rerun()
        st.info("Extract packing list data and generate shipment quote templates for carriers.")

    with col2:
        if st.button("🧾 Data Extractor & Paperwork Generator"):
            st.session_state.active_tool = "Invoice Extractor"
            st.rerun()
        st.info("Convert SAP Exports into formatted Customs Invoices with editable HTS summaries.")

# ==========================================
# 4. MAIN APP CONTENT
# ==========================================
else:
    if st.sidebar.button("⬅️ Back to Portal"):
        for key in ['df_detailed', 'ci_filled_bytes', 'vgw_filled_bytes',
                    'vgw_container_label', 'pd_filled_bytes', 'ausfta_filled_bytes']:
            if key in st.session_state: del st.session_state[key]
        st.session_state.active_tool = None
        st.rerun()

    # --- TOOL 1: QUOTE PIPELINE ---
    if st.session_state.active_tool == "Quote Pipeline":
        st.sidebar.title("Shipment Details")
        destinations = [
            "UK - Radial FAO Monat, 26, 26 Broadgate, Chadderton, Middleton Oldham OL9 9XA",
            "POLAND - Radial Poland Sp. z o.o. Moszna Parcela 29, Budynek C3 05-840 Brwinów",
            "AUSTRALIA - FDM WAREHOUSING C/O Landmark Global 7 Eucalyptus Place",
            "MONAT Global Canada — 135 SPARKS AVE NORTH YORK ON M2H 2S5 Canada",
            "FENIX FWD INC. - 417 LOGISTIC LAREDO, TEXAS 78045",
            "OTHER (Type Manually below)"
        ]
        services = ["40\" REEFER", "40\" DRY", "20\" DRY", "HAZMAT LCL", "LCL Ocean", "LTL Road", "Air Freight", "Courier"]

        selected_dest = st.sidebar.selectbox("Select Destination", destinations)
        destination = st.sidebar.text_input("Manual Destination Entry", value=selected_dest) if selected_dest == "OTHER (Type Manually below)" else selected_dest
        service = st.sidebar.selectbox("Service", services)
        commodity = st.sidebar.text_input("Commodity", value="Finished goods / Haircare / Skincare")
        cargo_value = st.sidebar.text_input("Value of Cargo", value="USD$ ")
        incoterms = st.sidebar.selectbox("Incoterms", ["-", "EXW", "FOB", "DDP", "DAP", "CIF"])

        st.title("📦 Quote Generator")
        packing_file = st.file_uploader("Upload Outbound Packing List (.xlsx)", type=['xlsx'])

        if packing_file:
            df_raw = pd.read_excel(packing_file, header=None).astype(str)

            def get_val(keyword, row_off=0, col_off=0):
                for r in range(len(df_raw)-1, -1, -1):
                    for c in range(len(df_raw.columns)):
                        if keyword.lower() == str(df_raw.iloc[r, c]).lower().strip():
                            try: return df_raw.iloc[r + row_off, c + col_off]
                            except: return "0"
                return "0"

            pallets_final = int(clean_numeric(get_val("Pallets", row_off=-1)))
            units_final   = int(clean_numeric(get_val("Units", row_off=-1)))
            lbs_final     = clean_numeric(get_val("Gross Weight", row_off=-1))
            kgs_final     = lbs_final * 0.453592

            dim_list = []
            for c in range(len(df_raw.columns)):
                if any("dim" in str(val).lower() and "pallet" in str(val).lower() for val in df_raw.iloc[:5, c]):
                    potential_dims = df_raw.iloc[3:, c].tolist()
                    dim_list = [d.strip() for d in potential_dims if "x" in str(d).lower() and len(str(d)) > 5]
                    break

            dim_counts    = Counter(dim_list)
            formatted_dims = [f"{d} (x{count})" if count > 1 else d for d, count in dim_counts.items()]

            st.success(f"✅ Data Extracted: **{pallets_final}** Pallets | **{units_final:,}** Units")

            if st.button("Generate Template"):
                quote_data = [
                    ["QUOTE REQUEST", ""], ["DESTINATION", destination], ["SERVICE", service],
                    ["UNITS", f"{units_final:,}"], ["PALLETS", pallets_final]
                ]
                if formatted_dims:
                    quote_data.append(["DIMENSIONS", formatted_dims[0]])
                    for extra_dim in formatted_dims[1:]: quote_data.append(["", extra_dim])

                quote_data.extend([["", ""], ["TOTAL WEIGHT", f"{lbs_final:,.2f} LBS | {kgs_final:,.2f} KGS"],
                                   ["COMMODITY", commodity], ["INCOTERMS", incoterms], ["VALUE OF CARGO", cargo_value]])

                df_output = pd.DataFrame(quote_data)
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                    df_output.to_excel(writer, index=False, header=False)

                dim_string = "".join([f"\n- • Dimensions: {d}" for d in formatted_dims])
                email_body = (
                    f"Hi Team,\n\nHope you are having a great week! \n\n"
                    f"Please find the details below for a new {service} shipment quote — please include insurance cost:\n\n"
                    f"- • Destination: {destination}\n- • Service: {service}\n"
                    f"- • Total Units: {units_final:,}\n- • Pallets: {pallets_final}{dim_string}\n"
                    f"- • Total Weight: {lbs_final:,.2f} LBS | {kgs_final:,.2f} KGS\n"
                    f"- • Commodity: {commodity}\n- • Value: {cargo_value}\n- • Incoterms: {incoterms}\n\n"
                    f"Thank you for your help."
                )

                st.divider()
                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("1. Download Document")
                    st.download_button("📥 Download Excel", data=buf.getvalue(), file_name=f"Quote_{pallets_final}PLTS.xlsx")
                    st.table(df_output)
                with col2:
                    st.subheader("2. Email Draft")
                    st.code(email_body, language="markdown")

    # --- TOOL 2: LOGISTICS PAPERWORK GENERATOR ---
    elif st.session_state.active_tool == "Invoice Extractor":
        st.title("🧾 Logistics Paperwork Generator")

        c1, c2 = st.columns(2)
        with c1: sap_file = st.file_uploader("1. Upload SAP Export", type=['csv', 'xlsx'])
        with c2: pl_file  = st.file_uploader("2. Upload Packing List", type=['csv', 'xlsx'])

        if sap_file and pl_file:
            hts_mapping = get_hts_data()

            # --- Robust Multi-PO Weight Extraction ---
            temp_pl = pd.read_excel(pl_file, header=None) if pl_file.name.endswith('.xlsx') else pd.read_csv(pl_file, header=None)
            weight_map   = {}
            current_cols = None

            for i, row in temp_pl.iterrows():
                row_vals = [str(x).strip() for x in row.values]
                if any(x in str(row_vals).lower() for x in ["sku", "material"]):
                    current_cols = [v.replace('\n', ' ').strip() for v in row_vals]
                    continue
                if current_cols:
                    row_dict = dict(zip(current_cols, row.values))
                    sku = clean_sku(row_dict.get('SKU') or row_dict.get('Material'))
                    if sku and sku != "nan":
                        tw_box  = clean_numeric(row_dict.get('Total Weight / Box') or row_dict.get('Tot. Weight / Bxs'))
                        t_units = clean_numeric(row_dict.get('Total Units'))
                        if t_units > 0:
                            weight_map[sku] = (tw_box / t_units) * 0.453592

            # --- SAP Processing ---
            if 'df_detailed' not in st.session_state:
                raw_df = pd.read_csv(sap_file) if sap_file.name.endswith('.csv') else pd.read_excel(sap_file)
                raw_df.columns = [str(col).strip() for col in raw_df.columns]

                rows = []
                for _, row in raw_df.iterrows():
                    sku = clean_sku(row.get('Material', ''))
                    if not sku: continue
                    sku_info = hts_mapping.get(sku, {"hts": "", "desc": ""})
                    qty      = clean_numeric(row.get('Order Quantity', 0))
                    u_price  = round(clean_numeric(row.get('Net Price', 0)) / 1000, 2)
                    u_weight = weight_map.get(sku, 0.0)
                    rows.append({
                        "PO#": str(row.get('Purchasing Document', '')),
                        "SKU": sku, "HTS Code": sku_info["hts"],
                        "Origin": "USA" if sku.startswith('600') else "CHINA" if sku.startswith('300') else "",
                        "Description": str(row.get('Short Text', '')).strip(),
                        "Quantity": int(qty), "Unit Price": u_price, "Total": round(qty * u_price, 2),
                        "Unit_Weight_KG": u_weight, "Total Weight (KG)": round(qty * u_weight, 2),
                        "Customs_Desc_Internal": sku_info["desc"]
                    })
                st.session_state.df_detailed = pd.DataFrame(rows)

            st.subheader("Commercial Invoice Data (Editable)")
            edited_detailed = st.data_editor(
                st.session_state.df_detailed.drop(columns=['Customs_Desc_Internal', 'Unit_Weight_KG']),
                use_container_width=True, hide_index=True,
                column_config={"Unit Price": st.column_config.NumberColumn(format="$%.3f")},
                key="detailed_editor", on_change=update_detailed_state
            )

            st.markdown("### HTS Summary (SLI Weight Breakdown)")
            summary_grouped = edited_detailed.merge(
                st.session_state.df_detailed[['SKU', 'Customs_Desc_Internal']], on='SKU', how='left'
            ).groupby(['Customs_Desc_Internal', 'HTS Code']).agg({
                'Quantity': 'sum', 'Total Weight (KG)': 'sum', 'Total': 'sum'
            }).reset_index()
            summary_grouped.columns = ['Customs Description', 'HTS Code', 'Total Qty', 'Total Weight (KG)', 'Total Value']

            st.data_editor(
                summary_grouped, use_container_width=True, hide_index=True,
                column_config={
                    "Total Value":        st.column_config.NumberColumn(format="$%.2f"),
                    "Total Weight (KG)":  st.column_config.NumberColumn(format="%.2f kg")
                },
                key="summary_editor"
            )

            # ── SLI Export + Commercial Invoice ───────────────────────────────
            st.divider()
            col_dl1, col_dl2 = st.columns(2)

            with col_dl1:
                st.subheader("📊 SLI Export")
                excel_buf = io.BytesIO()
                with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:
                    edited_detailed.to_excel(writer, index=False, sheet_name="Details")
                    summary_grouped.to_excel(writer, index=False, sheet_name="Summary")
                st.download_button(
                    "📥 Download SLI Excel", excel_buf.getvalue(), "SLI_Invoice.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            with col_dl2:
                st.subheader("📄 Commercial Invoice")
                selected_dest    = st.selectbox("Select Destination", options=list(INVOICE_DESTINATIONS.keys()), key="ci_destination")
                template_filename = INVOICE_DESTINATIONS[selected_dest]
                row_count        = len(st.session_state.df_detailed)

                if row_count > CI_MAX_ROWS:
                    st.warning(
                        f"⚠️ This shipment has **{row_count} SKUs**, which exceeds the "
                        f"**{CI_MAX_ROWS}-row** template capacity. Please fill this invoice manually."
                    )
                else:
                    if st.button("✍️ Fill Commercial Invoice Template"):
                        if 'ci_filled_bytes' in st.session_state: del st.session_state.ci_filled_bytes
                        ci_bytes, err = fill_commercial_invoice_template(st.session_state.df_detailed, template_filename)
                        if err: st.error(f"❌ {err}")
                        else:
                            st.session_state.ci_filled_bytes  = ci_bytes
                            st.session_state.ci_dest_label    = selected_dest
                            st.success("✅ Invoice filled! Click below to download.")

                    if 'ci_filled_bytes' in st.session_state:
                        po_str   = "_".join(str(p) for p in st.session_state.df_detailed['PO#'].dropna().unique())
                        dest_tag = template_filename.replace("Commercial_Invoice_", "").replace(".xlsx", "")
                        st.download_button(
                            "📥 Download Commercial Invoice",
                            data=st.session_state.ci_filled_bytes,
                            file_name=f"Commercial_Invoice_{dest_tag}_{po_str}_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            # ── VGW DECLARATION ───────────────────────────────────────────────
            st.divider()
            st.subheader("⚖️ Verified Gross Weight (VGW) Declaration")

            net_kg, gross_kg = extract_pl_weights_kg(temp_pl)

            vgw_c1, vgw_c2, vgw_c3 = st.columns(3)
            with vgw_c1: st.metric("Cargo Net Weight (KGS)",   f"{net_kg:,.2f}"   if net_kg   else "Not found")
            with vgw_c2: st.metric("Cargo Gross Weight (KGS)", f"{gross_kg:,.2f}" if gross_kg else "Not found")
            with vgw_c3: st.write("")

            vgw_m1, vgw_m2, vgw_m3 = st.columns(3)
            with vgw_m1: container_num = st.text_input("Container#",                    placeholder="e.g. HAMU 3039802", key="vgw_container")
            with vgw_m2: seal_num      = st.text_input("Seal#",                         placeholder="e.g. UL-9988229",   key="vgw_seal")
            with vgw_m3: tare_weight   = st.number_input("Container Tare Weight (KGS)", min_value=0.0, step=10.0,        key="vgw_tare")

            if st.button("✍️ Fill VGW Declaration"):
                if not container_num or not seal_num or tare_weight == 0.0:
                    st.warning("⚠️ Please fill in Container#, Seal#, and Tare Weight before generating.")
                elif net_kg == 0.0 or gross_kg == 0.0:
                    st.warning("⚠️ Could not extract Net/Gross Weight from the packing list. Please check the file.")
                else:
                    if 'vgw_filled_bytes' in st.session_state: del st.session_state.vgw_filled_bytes
                    vgw_bytes, err = fill_vgw_template(container_num, seal_num, tare_weight, net_kg, gross_kg)
                    if err: st.error(f"❌ {err}")
                    else:
                        st.session_state.vgw_filled_bytes      = vgw_bytes
                        st.session_state.vgw_container_label   = container_num
                        st.success("✅ VGW Declaration filled! Click below to download.")

            if 'vgw_filled_bytes' in st.session_state:
                container_tag = st.session_state.vgw_container_label.replace(" ", "_")
                st.download_button(
                    "📥 Download VGW Declaration",
                    data=st.session_state.vgw_filled_bytes,
                    file_name=f"VGW_Declaration_{container_tag}_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # ── APAC DOCUMENTS: PACKING DECLARATION + AUSFTA ─────────────────
            st.divider()
            st.subheader("🇦🇺 APAC Documents")

            apac_col1, apac_col2 = st.columns(2)

            # ── LEFT: Packing Declaration ─────────────────────────────────────
            with apac_col1:
                st.markdown("#### 📋 Packing Declaration")

                pd_m1, pd_m2 = st.columns(2)
                with pd_m1: pd_vessel      = st.text_input("Vessel Name",         placeholder="e.g. MAERSK SENTOSA",  key="pd_vessel")
                with pd_m2: pd_voyage      = st.text_input("Voyage Number",       placeholder="e.g. 123W",            key="pd_voyage")
                pd_consignment = st.text_input("Consignment Number (HBL/MBL/INV/CTNR)", placeholder="e.g. MSNU2598611", key="pd_consignment")
                pd_printed_name = st.text_input("Printed Name",                  placeholder="e.g. Kevin Alvarez",   key="pd_name")

                if st.button("✍️ Fill Packing Declaration"):
                    if not pd_vessel or not pd_voyage or not pd_consignment or not pd_printed_name:
                        st.warning("⚠️ Please fill in all fields before generating.")
                    else:
                        if 'pd_filled_bytes' in st.session_state: del st.session_state.pd_filled_bytes
                        pd_bytes, err = fill_packing_declaration_template(pd_vessel, pd_voyage, pd_consignment, pd_printed_name)
                        if err: st.error(f"❌ {err}")
                        else:
                            st.session_state.pd_filled_bytes       = pd_bytes
                            st.session_state.pd_consignment_label  = pd_consignment
                            st.success("✅ Packing Declaration filled! Click below to download.")

                if 'pd_filled_bytes' in st.session_state:
                    consignment_tag = st.session_state.pd_consignment_label.replace(" ", "_")
                    st.download_button(
                        "📥 Download Packing Declaration",
                        data=st.session_state.pd_filled_bytes,
                        file_name=f"Packing_Declaration_{consignment_tag}_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            # ── RIGHT: AUSFTA Declaration ─────────────────────────────────────
            with apac_col2:
                st.markdown("#### 🤝 AUSFTA Declaration")

                # PO# auto-pulled from SAP — shown as read-only info
                po_number = ", ".join(str(p) for p in st.session_state.df_detailed['PO#'].dropna().unique())
                st.info(f"📎 Invoice & Document Reference: **{po_number}** (from SAP export)")

                ausfta_bol  = st.text_input("Bill of Lading / Airway Bill", placeholder="e.g. S05336945/MISYD5336945", key="ausfta_bol")
                ausfta_name = st.text_input("Name Printed",                 placeholder="e.g. Kevin Alvarez",          key="ausfta_name")

                if st.button("✍️ Fill AUSFTA Declaration"):
                    if not ausfta_bol or not ausfta_name:
                        st.warning("⚠️ Please fill in Bill of Lading and Name Printed before generating.")
                    else:
                        if 'ausfta_filled_bytes' in st.session_state: del st.session_state.ausfta_filled_bytes
                        ausfta_bytes, err = fill_ausfta_template(ausfta_bol, ausfta_name, po_number)
                        if err: st.error(f"❌ {err}")
                        else:
                            st.session_state.ausfta_filled_bytes = ausfta_bytes
                            st.success("✅ AUSFTA Declaration filled! Click below to download.")

                if 'ausfta_filled_bytes' in st.session_state:
                    po_tag = po_number.replace(", ", "_")
                    st.download_button(
                        "📥 Download AUSFTA Declaration",
                        data=st.session_state.ausfta_filled_bytes,
                        file_name=f"AUSFTA_Declaration_{po_tag}_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
